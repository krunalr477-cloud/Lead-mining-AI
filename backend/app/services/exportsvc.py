"""Export service — materialize mining results to a downloadable file or Sheets.

``build_export(session, export_id)`` reads the tenant's DB-backed rows through
the same :mod:`app.sheetsync.tabs` TabSpecs the sheet-sync engine uses, so the
export is byte-for-byte consistent with the spreadsheet mirror and reuses the
sales-ready filter (VERIFIED + non-suppressed + non-tombstoned — spec §12 /
AC25) for free.

Scopes (spec §19 /exports body ``scope``):

- ``sales_ready`` — only the ``Sales_Ready_Leads`` tab (the clean output).
- ``raw`` — every DB-backed tab (the full mined dataset), shown separately from
  the clean output (spec §12 "show raw mined data separately").

Targets (spec §12 output, AC17):

- ``file`` — CSV / XLSX / JSON written under ``exports/``. For ``sales_ready``
  the CSV/JSON is a flat row list; ``raw`` produces a multi-sheet XLSX and a
  tab-keyed CSV/JSON.
- ``google_sheets`` — flush the tenant's spreadsheet via the existing sheet-sync
  engine (``run_sync``). No file is produced; the spreadsheet id is recorded as
  ``file_path``.

The ExportJob row's ``file_path``/``status``/``completed_at`` are updated in
place; the caller owns the transaction (the worker task commits).
"""

from __future__ import annotations

import csv
import json
import uuid
from pathlib import Path
from typing import Any

from sqlalchemy.orm import Session

from app.config import REPO_ROOT
from app.constants import ExportFormat, ExportScope, ExportTarget, JobStatus
from app.db import utcnow
from app.models import ExportJob
from app.sheetsync.tabs import TABS, TABS_BY_NAME, TabSpec

__all__ = ["EXPORTS_DIR", "build_export", "materialize_tabs"]

EXPORTS_DIR = REPO_ROOT / "exports"

# The single tab a sales_ready export projects.
SALES_READY_TAB = "Sales_Ready_Leads"


def _tabs_for_scope(scope: str) -> list[TabSpec]:
    """Which TabSpecs a scope materializes (README is static docs — skipped)."""
    if scope == ExportScope.SALES_READY:
        return [TABS_BY_NAME[SALES_READY_TAB]]
    return [t for t in TABS if t.name != "README"]


def materialize_tabs(
    session: Session, tenant_id: uuid.UUID, scope: str, job_id: uuid.UUID | None
) -> dict[str, dict[str, Any]]:
    """Build ``{tab_name: {"header": [...], "rows": [ {col: val} ]}}`` for a scope.

    Rows are projected through each ``TabSpec`` (same scalarization + sales-ready
    filter as the sheet-sync engine). When ``job_id`` is set, rows are filtered to
    that job wherever the tab carries a ``job_id`` column.
    """
    out: dict[str, dict[str, Any]] = {}
    for spec in _tabs_for_scope(scope):
        header = list(spec.columns)
        rows: list[dict[str, Any]] = []
        for raw_row in spec.source(session, tenant_id):
            if (
                job_id is not None
                and "job_id" in raw_row
                and str(raw_row.get("job_id") or "") != str(job_id)
            ):
                continue
            rows.append(spec.project(raw_row))  # scalarized, full-header dict
        out[spec.name] = {"header": header, "rows": rows}
    return out


def build_export(session: Session, export_id: uuid.UUID) -> dict:
    """Materialize an ExportJob to its target/format. Updates the row in place."""
    export = session.get(ExportJob, export_id)
    if export is None:
        return {"error": "export job not found"}

    export.status = JobStatus.RUNNING
    session.flush()
    try:
        if export.target == ExportTarget.GOOGLE_SHEETS:
            result = _export_to_sheets(session, export)
        else:
            result = _export_to_file(session, export)
        export.status = JobStatus.COMPLETED
        export.completed_at = utcnow()
        session.flush()
        return {"export_id": str(export.id), "status": export.status, **result}
    except Exception as exc:  # record the failure on the row and re-raise-safe
        export.status = JobStatus.FAILED
        export.error = str(exc)[:1000]
        export.completed_at = utcnow()
        session.flush()
        return {"export_id": str(export.id), "status": export.status, "error": str(exc)}


# --------------------------------------------------------------------------- #
# Google Sheets target — reuse the sheet-sync engine (no new Sheets code).
# --------------------------------------------------------------------------- #


def _export_to_sheets(session: Session, export: ExportJob) -> dict:
    from app.pipeline.stages import run_sync

    summary = run_sync(session, export.tenant_id)
    # Record the spreadsheet id as the "file_path" so GET /exports/{id} can point
    # the user at their sheet. run_sync stamps tenant.google_spreadsheet_id.
    from app.models import Tenant

    tenant = session.get(Tenant, export.tenant_id)
    export.file_path = tenant.google_spreadsheet_id if tenant else None
    return {"target": ExportTarget.GOOGLE_SHEETS, "spreadsheet_id": export.file_path, **summary}


# --------------------------------------------------------------------------- #
# File target — CSV / XLSX / JSON under exports/.
# --------------------------------------------------------------------------- #


def _export_to_file(session: Session, export: ExportJob) -> dict:
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    tabs = materialize_tabs(session, export.tenant_id, export.scope, export.job_id)
    fmt = export.format
    if fmt == ExportFormat.XLSX:
        path = _write_xlsx(export, tabs)
    elif fmt == ExportFormat.JSON:
        path = _write_json(export, tabs)
    else:
        path = _write_csv(export, tabs)
    export.file_path = str(path)
    row_count = sum(len(t["rows"]) for t in tabs.values())
    return {
        "target": ExportTarget.FILE,
        "format": fmt,
        "file": str(path),
        "tabs": list(tabs),
        "rows": row_count,
    }


def _is_flat(export: ExportJob) -> bool:
    """Sales-ready exports are a single flat tab; raw is multi-tab."""
    return export.scope == ExportScope.SALES_READY


def _write_csv(export: ExportJob, tabs: dict[str, dict[str, Any]]) -> Path:
    path = EXPORTS_DIR / f"export_{export.id}.csv"
    if _is_flat(export):
        tab = tabs.get(SALES_READY_TAB, {"header": [], "rows": []})
        _write_one_csv(path, tab["header"], tab["rows"])
        return path
    # Raw scope: prefix each row with its tab so a single CSV stays parseable,
    # unioning all headers. (Multi-file zips land with object storage — spec §6.)
    header: list[str] = ["_tab"]
    for tab in tabs.values():
        for col in tab["header"]:
            if col not in header:
                header.append(col)
    with path.open("w", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=header, extrasaction="ignore")
        writer.writeheader()
        for name, tab in tabs.items():
            for row in tab["rows"]:
                writer.writerow({"_tab": name, **row})
    return path


def _write_one_csv(path: Path, header: list[str], rows: list[dict]) -> None:
    with path.open("w", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=header, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({c: row.get(c, "") for c in header})


def _write_json(export: ExportJob, tabs: dict[str, dict[str, Any]]) -> Path:
    path = EXPORTS_DIR / f"export_{export.id}.json"
    if _is_flat(export):
        tab = tabs.get(SALES_READY_TAB, {"header": [], "rows": []})
        payload: Any = tab["rows"]
    else:
        payload = {name: tab["rows"] for name, tab in tabs.items()}
    path.write_text(json.dumps(payload, indent=2, default=str))
    return path


def _write_xlsx(export: ExportJob, tabs: dict[str, dict[str, Any]]) -> Path:
    from openpyxl import Workbook

    path = EXPORTS_DIR / f"export_{export.id}.xlsx"
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet
    if _is_flat(export):
        items = [(SALES_READY_TAB, tabs.get(SALES_READY_TAB, {"header": [], "rows": []}))]
    else:
        items = list(tabs.items())
    for name, tab in items:
        ws = wb.create_sheet(title=name[:31])  # Excel caps titles at 31 chars
        header = tab["header"]
        ws.append(list(header))
        ws.freeze_panes = "A2"
        if header:
            ws.auto_filter.ref = ws.dimensions or "A1"
        for row in tab["rows"]:
            ws.append([_xlsx_cell(row.get(col, "")) for col in header])
    if not wb.sheetnames:  # openpyxl refuses to save a book with zero sheets
        wb.create_sheet(title=SALES_READY_TAB[:31])
    wb.save(path)
    return path


def _xlsx_cell(value: Any) -> Any:
    """openpyxl accepts str/int/float/bool/None; coerce anything else to str."""
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)
