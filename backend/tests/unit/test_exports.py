"""Exports: CSV / XLSX / JSON materialization + scope filtering (spec §12, AC17).

Runs against the live Postgres schema via ``sync_session_factory`` (the workers'
sync engine), the same pattern as ``test_sheet_sync``. Each test builds an
isolated tenant and rolls back, leaving no residue. Files land in a temp dir so
we never touch the repo ``exports/``. No network: file/JSON/XLSX targets only;
the Google Sheets target is covered by the sheet-sync suite (it reuses the same
engine, driven here through the demo FakeSheetsClient path).
"""

from __future__ import annotations

import csv
import json
import uuid
from collections.abc import Iterator

import pytest
from sqlalchemy.orm import Session

from app.constants import (
    ExportFormat,
    ExportScope,
    ExportTarget,
    FinalEmailStatus,
    JobStatus,
)
from app.db import sync_session_factory
from app.models import ExportJob, SalesReadyLead, Tenant
from app.services import exportsvc


@pytest.fixture
def session() -> Iterator[Session]:
    s = sync_session_factory()
    try:
        yield s
    finally:
        s.rollback()
        s.close()


@pytest.fixture
def tenant(session: Session) -> Tenant:
    t = Tenant(name=f"export-test-{uuid.uuid4().hex[:8]}")
    session.add(t)
    session.flush()
    return t


@pytest.fixture(autouse=True)
def _exports_dir(tmp_path, monkeypatch) -> None:
    """Redirect the exports dir to a temp path so tests never write to the repo."""
    monkeypatch.setattr(exportsvc, "EXPORTS_DIR", tmp_path)


def _seed_lead(
    session: Session,
    tenant: Tenant,
    *,
    email: str,
    company: str,
    status: str = FinalEmailStatus.VERIFIED,
    tombstoned: bool = False,
    rank: int = 0,
    job_id: uuid.UUID | None = None,
) -> SalesReadyLead:
    lead = SalesReadyLead(
        tenant_id=tenant.id,
        job_id=job_id,
        company_name=company,
        email=email,
        contact_name="Ada Lovelace",
        designation="CTO",
        city="Austin",
        state="TX",
        country="US",
        services=["design", "dev"],
        validation_status=status,
        tombstoned=tombstoned,
        rank=rank,
    )
    session.add(lead)
    session.flush()
    return lead


def _make_export(
    session: Session,
    tenant: Tenant,
    fmt: str,
    *,
    scope: str = ExportScope.SALES_READY,
    target: str = ExportTarget.FILE,
    job_id: uuid.UUID | None = None,
) -> ExportJob:
    export = ExportJob(
        tenant_id=tenant.id,
        job_id=job_id,
        format=fmt,
        scope=scope,
        target=target,
        status=JobStatus.QUEUED,
    )
    session.add(export)
    session.flush()
    return export


# --------------------------------------------------------------------------- #
# CSV / XLSX / JSON sales-ready exports produce files with correct headers/rows
# --------------------------------------------------------------------------- #

SALES_READY_HEADER = [
    "sales_lead_id",
    "job_id",
    "company_name",
    "website",
    "city",
    "state",
    "country",
    "contact_name",
    "designation",
    "email",
    "phone",
    "services",
    "source_summary",
    "validation_status",
    "confidence_score",
    "last_verified_at",
    "campaign_status",
    "owner",
    "next_action",
    "sales_notes",
]


class TestSalesReadyFileExports:
    def test_csv(self, session: Session, tenant: Tenant) -> None:
        _seed_lead(session, tenant, email="ada@acme.example", company="Acme")
        _seed_lead(session, tenant, email="grace@beta.example", company="Beta")
        export = _make_export(session, tenant, ExportFormat.CSV)

        result = exportsvc.build_export(session, export.id)

        assert result["status"] == JobStatus.COMPLETED
        assert export.status == JobStatus.COMPLETED
        assert export.completed_at is not None
        from pathlib import Path

        path = Path(export.file_path)
        assert path.exists()
        with path.open(newline="") as fh:
            reader = csv.DictReader(fh)
            assert reader.fieldnames == SALES_READY_HEADER
            rows = list(reader)
        assert {r["email"] for r in rows} == {"ada@acme.example", "grace@beta.example"}
        assert {r["company_name"] for r in rows} == {"Acme", "Beta"}

    def test_json(self, session: Session, tenant: Tenant) -> None:
        _seed_lead(session, tenant, email="ada@acme.example", company="Acme")
        export = _make_export(session, tenant, ExportFormat.JSON)

        exportsvc.build_export(session, export.id)

        from pathlib import Path

        payload = json.loads(Path(export.file_path).read_text())
        assert isinstance(payload, list)  # flat row list for sales_ready
        assert len(payload) == 1
        assert payload[0]["email"] == "ada@acme.example"
        assert set(payload[0]) == set(SALES_READY_HEADER)

    def test_xlsx(self, session: Session, tenant: Tenant) -> None:
        _seed_lead(session, tenant, email="ada@acme.example", company="Acme")
        _seed_lead(session, tenant, email="grace@beta.example", company="Beta")
        export = _make_export(session, tenant, ExportFormat.XLSX)

        exportsvc.build_export(session, export.id)

        from pathlib import Path

        from openpyxl import load_workbook

        path = Path(export.file_path)
        assert path.exists()
        wb = load_workbook(path)
        assert wb.sheetnames == ["Sales_Ready_Leads"]
        ws = wb["Sales_Ready_Leads"]
        header = [c.value for c in ws[1]]
        assert header == SALES_READY_HEADER
        assert ws.freeze_panes == "A2"
        emails = {row[9].value for row in ws.iter_rows(min_row=2)}  # "email" col
        assert emails == {"ada@acme.example", "grace@beta.example"}


# --------------------------------------------------------------------------- #
# sales_ready scope excludes non-verified / tombstoned (spec §12 / AC25)
# --------------------------------------------------------------------------- #


class TestSalesReadyScopeFilters:
    def test_only_verified_non_tombstoned_appear(self, session: Session, tenant: Tenant) -> None:
        _seed_lead(session, tenant, email="good@acme.example", company="Acme")
        _seed_lead(
            session,
            tenant,
            email="role@beta.example",
            company="Beta",
            status=FinalEmailStatus.ROLE_BASED_REJECTED,
        )
        _seed_lead(
            session,
            tenant,
            email="invalid@gamma.example",
            company="Gamma",
            status=FinalEmailStatus.PROVIDER_INVALID,
        )
        _seed_lead(
            session,
            tenant,
            email="dead@delta.example",
            company="Delta",
            tombstoned=True,
        )
        export = _make_export(session, tenant, ExportFormat.JSON)

        exportsvc.build_export(session, export.id)

        from pathlib import Path

        payload = json.loads(Path(export.file_path).read_text())
        emails = {r["email"] for r in payload}
        assert emails == {"good@acme.example"}
        assert "role@beta.example" not in emails
        assert "invalid@gamma.example" not in emails
        assert "dead@delta.example" not in emails

    def test_materialize_tabs_sales_ready_is_single_tab(
        self, session: Session, tenant: Tenant
    ) -> None:
        _seed_lead(session, tenant, email="good@acme.example", company="Acme")
        tabs = exportsvc.materialize_tabs(session, tenant.id, ExportScope.SALES_READY, None)
        assert list(tabs) == ["Sales_Ready_Leads"]
        assert len(tabs["Sales_Ready_Leads"]["rows"]) == 1


# --------------------------------------------------------------------------- #
# raw scope spans every DB-backed tab (the full mined dataset)
# --------------------------------------------------------------------------- #


class TestRawScope:
    def test_raw_json_is_tab_keyed_and_multi_tab(self, session: Session, tenant: Tenant) -> None:
        _seed_lead(session, tenant, email="good@acme.example", company="Acme")
        export = _make_export(session, tenant, ExportFormat.JSON, scope=ExportScope.RAW)

        exportsvc.build_export(session, export.id)

        from pathlib import Path

        payload = json.loads(Path(export.file_path).read_text())
        assert isinstance(payload, dict)  # tab-keyed for raw
        # Every DB-backed tab is present (README is the only one skipped).
        assert "Sales_Ready_Leads" in payload
        assert "Companies" in payload
        assert "Contacts" in payload
        assert "README" not in payload
        assert len(payload["Sales_Ready_Leads"]) == 1

    def test_raw_xlsx_has_a_sheet_per_tab(self, session: Session, tenant: Tenant) -> None:
        _seed_lead(session, tenant, email="good@acme.example", company="Acme")
        export = _make_export(session, tenant, ExportFormat.XLSX, scope=ExportScope.RAW)

        exportsvc.build_export(session, export.id)

        from pathlib import Path

        from openpyxl import load_workbook

        wb = load_workbook(Path(export.file_path))
        assert "Sales_Ready_Leads" in wb.sheetnames
        assert "Companies" in wb.sheetnames
        assert "README" not in wb.sheetnames


# --------------------------------------------------------------------------- #
# job_id scoping
# --------------------------------------------------------------------------- #


class TestJobScoping:
    def test_job_id_filters_rows(self, session: Session, tenant: Tenant) -> None:
        from app.models import MiningJob

        job_a = MiningJob(tenant_id=tenant.id, name="Job A")
        job_b = MiningJob(tenant_id=tenant.id, name="Job B")
        session.add_all([job_a, job_b])
        session.flush()
        _seed_lead(session, tenant, email="a@acme.example", company="Acme", job_id=job_a.id)
        _seed_lead(session, tenant, email="b@beta.example", company="Beta", job_id=job_b.id)
        tabs = exportsvc.materialize_tabs(session, tenant.id, ExportScope.SALES_READY, job_a.id)
        emails = {r["email"] for r in tabs["Sales_Ready_Leads"]["rows"]}
        assert emails == {"a@acme.example"}


# --------------------------------------------------------------------------- #
# failure path records the error on the row
# --------------------------------------------------------------------------- #


def test_missing_export_returns_error(session: Session) -> None:
    result = exportsvc.build_export(session, uuid.uuid4())
    assert result == {"error": "export job not found"}
