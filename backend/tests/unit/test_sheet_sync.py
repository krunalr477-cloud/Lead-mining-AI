"""Sheet sync engine: idempotency, editable-field protection, tab contracts.

Runs against the live Postgres schema via ``sync_session_factory`` (workers'
engine). Each test builds an isolated tenant and rolls back at the end, so runs
leave no residue. The FakeSheetsClient is constructed with ``persist=False`` in
most tests to avoid touching ``exports/``; persistence and .xlsx dump have their
own dedicated tests.
"""

from __future__ import annotations

import uuid
from collections.abc import Iterator
from decimal import Decimal

import pytest
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.constants import (
    EnrichmentStatus,
    FinalEmailStatus,
    StageStatus,
    SyncStatus,
)
from app.db import sync_session_factory, utcnow
from app.models import (
    Company,
    Contact,
    EmailCandidate,
    SalesReadyLead,
    SheetRowMap,
    Tenant,
    ValidationCheck,
)
from app.sheetsync import FakeSheetsClient, SheetSyncEngine
from app.sheetsync.engine import CYAN, GREEN, RED, STATUS_COLORS
from app.sheetsync.tabs import TABS, TABS_BY_NAME


@pytest.fixture
def session() -> Iterator[Session]:
    s = sync_session_factory()
    try:
        yield s
    finally:
        s.rollback()
        s.close()


@pytest.fixture
def tenant(session: Session) -> Tenant:
    t = Tenant(name=f"sheet-sync-test-{uuid.uuid4().hex[:8]}")
    session.add(t)
    session.flush()
    return t


def _seed_company(session: Session, tenant: Tenant, name: str) -> Company:
    company = Company(
        tenant_id=tenant.id,
        canonical_name=name,
        website=f"https://{name.lower()}.example",
        domain=f"{name.lower()}.example",
        city="Austin",
        state="TX",
        country="US",
        services=["design", "dev"],
        compliance_posture="green",
    )
    session.add(company)
    session.flush()
    return company


def _seed_contact(
    session: Session,
    tenant: Tenant,
    company: Company,
    *,
    email: str,
    final_status: str | None,
) -> Contact:
    contact = Contact(
        tenant_id=tenant.id,
        company_id=company.id,
        full_name="Ada Lovelace",
        first_name="Ada",
        last_name="Lovelace",
        designation="CTO",
        email=email,
        enrichment_status=EnrichmentStatus.ENRICHED,
        final_email_status=final_status,
        confidence_score=Decimal("0.910"),
        sales_ready=final_status == FinalEmailStatus.VERIFIED,
        notes="original sales note",
    )
    session.add(contact)
    session.flush()
    return contact


def _seed_validation(
    session: Session, contact: Contact, company: Company, *, final: str
) -> ValidationCheck:
    cand = EmailCandidate(contact_id=contact.id, email=contact.email, source="pattern")
    session.add(cand)
    session.flush()
    check = ValidationCheck(
        email_candidate_id=cand.id,
        contact_id=contact.id,
        company_id=company.id,
        syntax_status=StageStatus.PASS,
        disposable_status=StageStatus.PASS,
        role_based_status=StageStatus.PASS,
        mx_status=StageStatus.PASS,
        final_status=final,
        verified_at=utcnow(),
    )
    session.add(check)
    session.flush()
    return check


# --------------------------------------------------------------------------- #


class TestTabContracts:
    def test_twelve_tabs_in_order(self) -> None:
        names = [t.name for t in TABS]
        assert names == [
            "README",
            "Mining_Jobs",
            "Companies",
            "Contacts",
            "Email_Validation",
            "Sales_Ready_Leads",
            "Outreach_Queue",
            "Campaigns",
            "Bounce_Log",
            "Suppression_List",
            "Data_Source_Audit",
            "Audit_Log",
        ]

    def test_editable_columns_only_on_contacts_and_sales_ready(self) -> None:
        editable = {t.name: set(t.editable_columns) for t in TABS}
        expected = {"owner", "sales_notes", "next_action"}
        assert editable["Contacts"] == expected
        assert editable["Sales_Ready_Leads"] == expected
        for name, cols in editable.items():
            if name not in ("Contacts", "Sales_Ready_Leads"):
                assert cols == set(), f"{name} must have no editable columns"

    def test_system_columns_exclude_editable(self) -> None:
        contacts = TABS_BY_NAME["Contacts"]
        for col in contacts.editable_columns:
            assert col not in contacts.system_columns()

    def test_key_columns(self) -> None:
        assert TABS_BY_NAME["Contacts"].key_column == "contact_id"
        assert TABS_BY_NAME["Sales_Ready_Leads"].key_column == "sales_lead_id"
        assert TABS_BY_NAME["Companies"].key_column == "company_id"


class TestSetupAndFormatting:
    def test_setup_creates_all_tabs_with_headers(self, session: Session, tenant: Tenant) -> None:
        client = FakeSheetsClient(persist=False)
        engine = SheetSyncEngine(session, client)
        engine.setup_spreadsheet(tenant.id)
        assert set(client.tabs) == {t.name for t in TABS}
        for t in TABS:
            assert client.tabs[t.name]["header"] == list(t.columns)

    def test_formatting_records_freeze_filter_and_colors(
        self, session: Session, tenant: Tenant
    ) -> None:
        client = FakeSheetsClient(persist=False)
        engine = SheetSyncEngine(session, client)
        engine.setup_spreadsheet(tenant.id)
        # Every tab freezes its header and enables a filter.
        for t in TABS:
            fmt = client.formatting[t.name]
            assert fmt.freeze_header is True
            assert fmt.filter_enabled is True
        # A tab with status columns records the status->color palette.
        contacts_fmt = client.formatting["Contacts"]
        assert "final_email_status" in contacts_fmt.status_columns
        assert contacts_fmt.status_colors[FinalEmailStatus.VERIFIED] == GREEN

    def test_status_color_palette_buckets(self) -> None:
        # Spec §5 lines 452-456: green/red/amber-purple/cyan buckets all present.
        assert STATUS_COLORS[FinalEmailStatus.VERIFIED] == GREEN
        assert STATUS_COLORS["valid"] == GREEN
        assert STATUS_COLORS["delivered"] == GREEN
        assert STATUS_COLORS[FinalEmailStatus.INVALID_SYNTAX] == RED
        assert STATUS_COLORS["hard_bounce"] == RED
        assert STATUS_COLORS["catch_all"] in ("#FFB020", "#A66BFF")  # amber/purple
        assert STATUS_COLORS["risk"] == "#A66BFF"
        assert STATUS_COLORS["running"] == CYAN
        assert STATUS_COLORS["queued"] == CYAN


class TestIdempotency:
    def test_second_flush_is_a_noop(self, session: Session, tenant: Tenant) -> None:
        company = _seed_company(session, tenant, "Acme")
        _seed_contact(
            session,
            tenant,
            company,
            email="ada@acme.example",
            final_status=FinalEmailStatus.VERIFIED,
        )
        _seed_contact(
            session,
            tenant,
            company,
            email="grace@acme.example",
            final_status=FinalEmailStatus.CATCH_ALL_REVIEW,
        )

        client = FakeSheetsClient(persist=False)
        engine = SheetSyncEngine(session, client)
        engine.setup_spreadsheet(tenant.id)

        first = engine.flush_tab(tenant.id, "Contacts")
        assert first.appended == 2
        assert first.updated == 0
        assert client.append_count == 2

        # Second flush: no DB change ⇒ zero writes.
        client.append_count = 0
        client.update_count = 0
        second = engine.flush_tab(tenant.id, "Contacts")
        assert second.writes == 0
        assert second.appended == 0
        assert second.updated == 0
        assert second.skipped == 2
        assert client.append_count == 0
        assert client.update_count == 0

    def test_row_map_records_row_numbers(self, session: Session, tenant: Tenant) -> None:
        company = _seed_company(session, tenant, "Acme")
        _seed_contact(
            session,
            tenant,
            company,
            email="ada@acme.example",
            final_status=FinalEmailStatus.VERIFIED,
        )
        client = FakeSheetsClient(persist=False)
        engine = SheetSyncEngine(session, client)
        engine.setup_spreadsheet(tenant.id)
        engine.flush_tab(tenant.id, "Contacts")
        maps = session.scalars(select(SheetRowMap).where(SheetRowMap.tenant_id == tenant.id)).all()
        # First data row lands at sheet row 2 (header is row 1).
        contact_maps = [m for m in maps if m.tab == "Contacts"]
        assert len(contact_maps) == 1
        assert contact_maps[0].row_number == 2


class TestSystemFieldUpdate:
    def test_system_change_writes_exactly_that_row(self, session: Session, tenant: Tenant) -> None:
        company = _seed_company(session, tenant, "Acme")
        c1 = _seed_contact(
            session,
            tenant,
            company,
            email="ada@acme.example",
            final_status=FinalEmailStatus.CATCH_ALL_REVIEW,
        )
        _seed_contact(
            session,
            tenant,
            company,
            email="grace@acme.example",
            final_status=FinalEmailStatus.VERIFIED,
        )
        client = FakeSheetsClient(persist=False)
        engine = SheetSyncEngine(session, client)
        engine.setup_spreadsheet(tenant.id)
        engine.flush_tab(tenant.id, "Contacts")

        # Change a SYSTEM column (final_email_status) on exactly one contact.
        c1.final_email_status = FinalEmailStatus.VERIFIED
        c1.sales_ready = True
        session.flush()

        client.append_count = 0
        client.update_count = 0
        result = engine.flush_tab(tenant.id, "Contacts")
        assert result.appended == 0
        assert result.updated == 1  # exactly one row rewritten
        assert result.skipped == 1
        assert client.update_count == 1
        # The sheet cell for that row now reflects the new system value.
        rows = client.tabs["Contacts"]["rows"]
        changed = [r for r in rows if r["contact_id"] == str(c1.id)][0]
        assert changed["final_email_status"] == FinalEmailStatus.VERIFIED


class TestEditableProtection:
    def test_editable_db_change_does_not_write(self, session: Session, tenant: Tenant) -> None:
        company = _seed_company(session, tenant, "Acme")
        contact = _seed_contact(
            session,
            tenant,
            company,
            email="ada@acme.example",
            final_status=FinalEmailStatus.VERIFIED,
        )
        client = FakeSheetsClient(persist=False)
        engine = SheetSyncEngine(session, client)
        engine.setup_spreadsheet(tenant.id)
        engine.flush_tab(tenant.id, "Contacts")

        # Change an EDITABLE field in the DB source (notes -> sales_notes column).
        contact.notes = "sales edited this in the DB"
        session.flush()

        client.append_count = 0
        client.update_count = 0
        result = engine.flush_tab(tenant.id, "Contacts")
        # Editable change is invisible to the engine: zero writes.
        assert result.writes == 0
        assert result.updated == 0
        assert client.update_count == 0

    def test_update_payload_never_contains_editable_columns(
        self, session: Session, tenant: Tenant
    ) -> None:
        """When a system field changes, the pushed range excludes editable cells.

        A sales user's edit lives only in the sheet; we prove the engine's update
        payload for a changed row contains no editable column, so the client can
        never overwrite it.
        """
        company = _seed_company(session, tenant, "Acme")
        contact = _seed_contact(
            session,
            tenant,
            company,
            email="ada@acme.example",
            final_status=FinalEmailStatus.CATCH_ALL_REVIEW,
        )
        client = FakeSheetsClient(persist=False)
        engine = SheetSyncEngine(session, client)
        engine.setup_spreadsheet(tenant.id)
        engine.flush_tab(tenant.id, "Contacts")

        # Simulate a sales edit that exists ONLY in the sheet mirror.
        sheet_row = client.tabs["Contacts"]["rows"][0]
        sheet_row["owner"] = "sales-user-owned"
        sheet_row["sales_notes"] = "do not clobber me"
        sheet_row["next_action"] = "call friday"

        # Now change a system field in the DB and flush.
        contact.final_email_status = FinalEmailStatus.VERIFIED
        session.flush()

        captured: list[dict] = []
        orig_update = client.update_ranges

        def spy(tab, header, updates):  # type: ignore[no-untyped-def]
            for u in updates:
                captured.append(u.columns)
            return orig_update(tab, header, updates)

        client.update_ranges = spy  # type: ignore[method-assign]
        result = engine.flush_tab(tenant.id, "Contacts")
        assert result.updated == 1
        assert captured, "expected an update payload"
        editable = set(TABS_BY_NAME["Contacts"].editable_columns)
        for cols in captured:
            assert not (set(cols) & editable), "update pushed an editable column!"

        # The sales edits survived untouched in the sheet mirror.
        assert sheet_row["owner"] == "sales-user-owned"
        assert sheet_row["sales_notes"] == "do not clobber me"
        assert sheet_row["next_action"] == "call friday"
        # But the system column WAS updated.
        assert sheet_row["final_email_status"] == FinalEmailStatus.VERIFIED


class TestSalesReadyFilter:
    def test_only_verified_leads_appear(self, session: Session, tenant: Tenant) -> None:
        # One VERIFIED lead, one non-verified, one tombstoned VERIFIED.
        good = SalesReadyLead(
            tenant_id=tenant.id,
            company_name="Acme",
            email="ada@acme.example",
            validation_status=FinalEmailStatus.VERIFIED,
            rank=1,
        )
        bad = SalesReadyLead(
            tenant_id=tenant.id,
            company_name="Beta",
            email="role@beta.example",
            validation_status=FinalEmailStatus.ROLE_BASED_REJECTED,
            rank=2,
        )
        tombstoned = SalesReadyLead(
            tenant_id=tenant.id,
            company_name="Gamma",
            email="x@gamma.example",
            validation_status=FinalEmailStatus.VERIFIED,
            tombstoned=True,
            rank=3,
        )
        session.add_all([good, bad, tombstoned])
        session.flush()

        client = FakeSheetsClient(persist=False)
        engine = SheetSyncEngine(session, client)
        engine.setup_spreadsheet(tenant.id)
        result = engine.flush_tab(tenant.id, "Sales_Ready_Leads")

        assert result.appended == 1  # only the verified, non-tombstoned lead
        rows = client.tabs["Sales_Ready_Leads"]["rows"]
        emails = {r["email"] for r in rows}
        assert emails == {"ada@acme.example"}
        assert "role@beta.example" not in emails
        assert "x@gamma.example" not in emails


class TestSyncEventDraining:
    def test_pending_events_marked_synced(self, session: Session, tenant: Tenant) -> None:
        company = _seed_company(session, tenant, "Acme")
        contact = _seed_contact(
            session,
            tenant,
            company,
            email="ada@acme.example",
            final_status=FinalEmailStatus.VERIFIED,
        )
        client = FakeSheetsClient(persist=False)
        engine = SheetSyncEngine(session, client)
        engine.setup_spreadsheet(tenant.id)

        ev = engine.enqueue_upsert(session, tenant.id, "Contacts", str(contact.id))
        assert ev.status == SyncStatus.PENDING

        result = engine.flush_tab(tenant.id, "Contacts")
        assert result.events_synced == 1
        session.refresh(ev)
        assert ev.status == SyncStatus.SYNCED
        assert ev.synced_at is not None

    def test_enqueue_rejects_unknown_tab(self, session: Session, tenant: Tenant) -> None:
        engine = SheetSyncEngine(session, FakeSheetsClient(persist=False))
        with pytest.raises(ValueError):
            engine.enqueue_upsert(session, tenant.id, "Not_A_Tab", "k")


class TestPersistenceAndExport:
    def test_json_mirror_and_xlsx_dump(self, session: Session, tenant: Tenant) -> None:
        company = _seed_company(session, tenant, "Acme")
        _seed_contact(
            session,
            tenant,
            company,
            email="ada@acme.example",
            final_status=FinalEmailStatus.VERIFIED,
        )
        client = FakeSheetsClient(tenant.id, persist=True)
        engine = SheetSyncEngine(session, client)
        engine.setup_spreadsheet(tenant.id)
        engine.flush_tab(tenant.id, "Contacts")

        mirror = client.mirror_path()
        assert mirror.exists()

        xlsx = client.dump_xlsx()
        try:
            assert xlsx.exists()
            from openpyxl import load_workbook

            wb = load_workbook(xlsx)
            assert "Contacts" in wb.sheetnames
            ws = wb["Contacts"]
            assert ws.freeze_panes == "A2"  # header frozen
            header = [c.value for c in ws[1]]
            assert header == list(TABS_BY_NAME["Contacts"].columns)
        finally:
            xlsx.unlink(missing_ok=True)
            mirror.unlink(missing_ok=True)


def test_google_client_stub_raises() -> None:
    from app.sheetsync import GoogleSheetsClient

    client = GoogleSheetsClient()
    with pytest.raises(NotImplementedError, match="Phase 3"):
        client.ensure_spreadsheet(uuid.uuid4())
